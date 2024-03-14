import openpyxl
from openpyxl.styles import PatternFill

def refine_excel(save_excel_file_path):
    wb = openpyxl.load_workbook(save_excel_file_path)
    ws = wb['all']
    for col_idx, col in enumerate(ws.iter_cols()):
        if col_idx == 0:
            pass
        else:
            col_name = col[0].value
            if 'dist' in col_name:
                row_judge_val = min([cell.value for cell in col if type(cell.value) is not str])
            else:
                row_judge_val = max([cell.value for cell in col if type(cell.value) is not str])
            
            for row_idx, cell in enumerate(col):
                cell.number_format = '0.0'
                if cell.value == row_judge_val:
                    cell.fill = PatternFill(fgColor='FFFF00', bgColor="FFFF00", fill_type = "solid")
                    cell.value = r'\red' + '{' + f'{cell.value:.1f}' + '}'
                # cell.fill = PatternFill(fgColor='FFFF00', bgColor="FFFF00", fill_type = "solid")
                # cell.value = r'\red' + '{' + f'{cell.value:.3f}' + '}'

    wb.save(save_excel_file_path)